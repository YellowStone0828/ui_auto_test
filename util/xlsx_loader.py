# -*- coding: utf-8 -*-

' xlsx load module '

__author__ = 'Jack Sun'

from openpyxl import load_workbook

# 按路径来解析excel
def xlsxProsser(filePath, sheetName):

    wb = load_workbook(filename=filePath)
    # wb = load_workbook(filename='./cases/test/double.xlsx')

    # 获取所有表格(worksheet)的名字
    # sheets = wb.get_sheet_names()

    # 获取特定的 worksheet
    ws = wb.get_sheet_by_name(sheetName)

    # 获取表格所有行和列
    rows = ws.rows
    # columns = ws.columns

    # 通过坐标读取值
    # ws.cell('B12').value  # B 表示列，12 表示行
    # ws.cell(row=12, column=2).value

    # 行循环解析
    content = []
    result = []

    for row in rows:
        line = [col.value for col in row]
        content.append(line)

    # 忽略表格名和第一行定义
    content.pop(0)
    idContent = content[0]
    content.pop(0)

    # 循环解析对象并返回
    for eachAction in content:
        actionDetail = {}
        checkNoneNum = 0
        for Index in range(len(eachAction)):
            actionDetail[idContent[Index]] = eachAction[Index]
        for eachAtr in actionDetail:
            if actionDetail[eachAtr] is None:
                checkNoneNum = checkNoneNum + 1
        if checkNoneNum != len(actionDetail):
            result.append(actionDetail)
    # print(result)
    return result


if __name__ == '__main__':
    xlsxProsser('../cases/test/double.xlsx', 'case')
